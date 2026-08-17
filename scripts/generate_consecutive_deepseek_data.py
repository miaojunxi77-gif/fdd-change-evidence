#!/usr/bin/env python3
"""Build website-ready consecutive-year shards from the cleaned DeepSeek workbook.

The workbook's analysis unit is an atomic clause change.  The website's evidence
browser uses one row per brand–Item–year comparison, so this script groups all
included conservative atomic changes belonging to the same job while preserving
their evidence quotations and page locators.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


MODEL = "deepseek-v4-flash-0731"
MAX_SHARD_BYTES = 700_000
CONTRACTUAL_BINDING = {
    "contractual_unconditional",
    "contractual_conditional",
    "franchisor_discretion",
}
IGNORED_DIRECTIONS = {None, "", "none", "unclear", "not_applicable", "not_comparable", "stable"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output_dir", type=Path)
    return parser.parse_args()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def shorten(value: Any, limit: int = 240) -> str:
    text = clean_text(value)
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "…"


def parse_evidence(value: Any) -> list[dict[str, Any]]:
    if value in (None, ""):
        return []
    if isinstance(value, str):
        try:
            payload = json.loads(value)
        except json.JSONDecodeError:
            payload = [{"quote": value, "page": None}]
    elif isinstance(value, list):
        payload = value
    else:
        payload = [{"quote": str(value), "page": None}]

    result: list[dict[str, Any]] = []
    seen: set[tuple[str, int | None]] = set()
    for entry in payload:
        if not isinstance(entry, dict):
            entry = {"quote": str(entry), "page": None}
        quote = clean_text(entry.get("quote"))
        page_raw = entry.get("page")
        try:
            page = int(page_raw) if page_raw not in (None, "") else None
        except (TypeError, ValueError):
            page = None
        if not quote or (quote, page) in seen:
            continue
        seen.add((quote, page))
        result.append(
            {
                "quote": quote,
                "page": page,
                "claimedPage": page,
                "verified": False,
                "method": "deepseek_cleaned_export",
            }
        )
    return result


def unique_evidence(rows: Iterable[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[tuple[str, int | None]] = set()
    for row in rows:
        for entry in parse_evidence(row.get(key)):
            marker = (entry["quote"], entry["page"])
            if marker not in seen:
                seen.add(marker)
                result.append(entry)
    return result


def page_range(evidence: list[dict[str, Any]]) -> list[int | None]:
    pages = [entry["page"] for entry in evidence if entry["page"] is not None]
    return [min(pages), max(pages)] if pages else [None, None]


def pair_id(job_id: str) -> str:
    value = re.sub(r"^fdd__consecutive__", "", job_id)
    return re.sub(r"__item_\d+$", "", value)


def select_direction(rows: list[dict[str, Any]]) -> str:
    for key in ("final_economic_burden_direction", "final_control_shift", "final_numeric_direction"):
        values = {clean_text(row.get(key)) for row in rows}
        values -= IGNORED_DIRECTIONS
        if len(values) == 1:
            return next(iter(values))
        if len(values) > 1:
            return "mixed"
    changes = {clean_text(row.get("change_type")) for row in rows if clean_text(row.get("change_type"))}
    return next(iter(changes)) if len(changes) == 1 else "mixed"


def build_summary(rows: list[dict[str, Any]]) -> str:
    ordered = sorted(rows, key=lambda row: (-int(row.get("score_numeric") or 0), clean_text(row.get("variable_name"))))
    snippets: list[str] = []
    for row in ordered:
        reason = shorten(row.get("score_reason"), 220)
        variable = shorten(row.get("variable_name"), 90)
        text = f"{variable}: {reason}" if variable and reason else reason or variable
        if text and text not in snippets:
            snippets.append(text)
        if len(snippets) == 3:
            break
    lead = f"{len(rows)} conservative atomic change{'s' if len(rows) != 1 else ''}."
    remainder = len(rows) - len(snippets)
    body = " ".join(snippets)
    tail = f" Plus {remainder} additional included change{'s' if remainder != 1 else ''}." if remainder > 0 else ""
    return clean_text(f"{lead} {body}{tail}")


def build_inference_limit(rows: list[dict[str, Any]]) -> str:
    values: list[str] = []
    for row in rows:
        value = clean_text(row.get("inference_limit"))
        if value and value not in values:
            values.append(value)
        if len(values) == 3:
            break
    return " ".join(values)


def read_sheet_rows(workbook: Path, sheet: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[sheet]
    iterator = ws.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(iterator)]
    return [dict(zip(headers, values)) for values in iterator if any(value is not None for value in values)]


def main() -> None:
    args = parse_args()
    atomic_rows = read_sheet_rows(args.workbook, "Clean Atomic")
    item_rows = read_sheet_rows(args.workbook, "Item Results")
    item_titles = {int(row["item"]): clean_text(row["item_title"]) for row in item_rows}

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in atomic_rows:
        if clean_text(row.get("final_status")) != "INCLUDE_CONSERVATIVE":
            continue
        groups[clean_text(row["job_id"])].append(row)

    website_rows: list[dict[str, Any]] = []
    for job_id, rows in sorted(groups.items()):
        first = rows[0]
        old_evidence = unique_evidence(rows, "old_evidence")
        new_evidence = unique_evidence(rows, "new_evidence")
        score = max(int(row.get("score_numeric") or 0) for row in rows)
        contractual = any(
            int(row.get("score_numeric") or 0) >= 4
            and clean_text(row.get("binding_status")) in CONTRACTUAL_BINDING
            for row in rows
        )
        manual_notes = [clean_text(row.get("manual_note")) for row in rows if clean_text(row.get("manual_note"))]
        item = int(first["item"])
        website_rows.append(
            {
                "route": "consecutive",
                "id": job_id,
                "pairId": pair_id(job_id),
                "company": clean_text(first.get("company")),
                "oldYear": int(first["old_year"]),
                "newYear": int(first["new_year"]),
                "yearGap": int(first["new_year"]) - int(first["old_year"]),
                "oldSource": "DeepSeek cleaned export",
                "newSource": "DeepSeek cleaned export",
                "sameSource": True,
                "item": item,
                "itemTitle": item_titles[item],
                "score": score,
                "substantive": True,
                "contractual": contractual,
                "routine": False,
                "direction": select_direction(rows),
                "summary": build_summary(rows),
                "statedReason": "",
                "confidence": "analysis-ready",
                "needsReview": False,
                "reviewReason": " ".join(dict.fromkeys(manual_notes)),
                "evidenceStatus": "deepseek_conservative_cleaned",
                "evidenceGatePass": True,
                "oldDocument": "",
                "newDocument": "",
                "oldPages": page_range(old_evidence),
                "newPages": page_range(new_evidence),
                "oldEvidence": old_evidence,
                "newEvidence": new_evidence,
                "samplingWeight": None,
                "samplingStratum": "",
                "inSbaDirectory": False,
                "replacement": False,
                "replacementLevel": "",
                "financingGuardrailPass": item != 10 or all(clean_text(row.get("financing_type")) != "unsupported" for row in rows),
                "financingWarnings": "",
                "scoringSource": "deepseek_api_cleaned",
                "model": MODEL,
                "atomicChangeCount": len(rows),
                "highImpactAtomicChanges": sum(int(row.get("score_numeric") or 0) >= 4 for row in rows),
                "inferenceLimit": build_inference_limit(rows),
            }
        )

    by_item: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in website_rows:
        by_item[row["item"]].append(row)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    filenames: list[str] = []
    for item in range(1, 24):
        filename = f"item-{item:02d}.json"
        rows = by_item.get(item, [])
        payload = json.dumps({"rows": rows}, ensure_ascii=False, separators=(",", ":"))
        if len(payload.encode("utf-8")) <= MAX_SHARD_BYTES:
            filenames.append(filename)
            (args.output_dir / filename).write_text(payload, encoding="utf-8")
            continue

        midpoint = (len(rows) + 1) // 2
        chunks = (rows[:midpoint], rows[midpoint:])
        for index, chunk in enumerate(chunks):
            chunk_filename = filename if index == 0 else f"item-{item:02d}-b.json"
            filenames.append(chunk_filename)
            (args.output_dir / chunk_filename).write_text(
                json.dumps({"rows": chunk}, ensure_ascii=False, separators=(",", ":")),
                encoding="utf-8",
            )

    metadata = {
        "route": "consecutive",
        "design": "DeepSeek v4.3 conservative cleaned consecutive-year results",
        "model": MODEL,
        "preparedJobs": 4557,
        "exportedJobs": 4556,
        "completeComparisons": 4352,
        "comparisons": len(website_rows),
        "includedChangeJobs": len(website_rows),
        "atomicChanges": len(atomic_rows),
        "highImpactChangeJobs": sum(row["score"] >= 4 for row in website_rows),
        "highImpactAtomicChanges": sum(row["highImpactAtomicChanges"] for row in website_rows),
        "uniquePairs": len({row["pairId"] for row in website_rows}),
        "items": list(range(1, 24)),
        "rowsPerItem": {str(item): len(by_item.get(item, [])) for item in range(1, 24)},
        "includesNoChange": False,
        "unit": "included_change_job",
        "source": args.workbook.name,
        "note": "Row-level explorer contains analysis-ready comparisons with at least one included conservative change; denominator counts and no-change jobs are reported in the aggregate tables.",
    }
    (args.output_dir / "index.json").write_text(
        json.dumps({"metadata": metadata, "files": filenames}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    assert len(website_rows) == 1661, len(website_rows)
    assert len(atomic_rows) == 7705, len(atomic_rows)
    assert metadata["highImpactChangeJobs"] == 1080, metadata["highImpactChangeJobs"]
    assert metadata["highImpactAtomicChanges"] == 3282, metadata["highImpactAtomicChanges"]

    print(json.dumps(metadata, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
