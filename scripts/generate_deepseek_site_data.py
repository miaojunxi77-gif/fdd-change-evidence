#!/usr/bin/env python3
"""Generate the public DeepSeek evidence datasets used by the website.

The consecutive workbook contains only final conservative atomic changes.  The
cross-period package contains every successful comparison plus candidate atomic
changes, outcome-readiness flags, and review flags.  This generator preserves
those distinctions and exposes one consistent four-category display taxonomy.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


MODEL = "deepseek-v4-flash-0731"
MAX_SHARD_BYTES = 360_000
DISPLAY_CHANGE_TYPES = ("introduced", "modified", "removed", "reclassified")
CHANGE_TYPE_LABELS = {
    "introduced": "新增",
    "modified": "修改",
    "removed": "删除",
    "reclassified": "重新分类",
}
CONTRACTUAL_BINDING = {
    "contractual_unconditional",
    "contractual_conditional",
    "franchisor_discretion",
}
IGNORED_DIRECTIONS = {"", "none", "unclear", "not_applicable", "not_comparable", "stable"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("consecutive_workbook", type=Path)
    parser.add_argument("cross_package_dir", type=Path)
    parser.add_argument("site_root", type=Path)
    return parser.parse_args()


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return clean(value).lower() in {"true", "1", "yes"}


def as_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def parse_json(value: Any, default: Any) -> Any:
    if value in (None, ""):
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except (json.JSONDecodeError, TypeError):
        return default


def read_sheet_rows(workbook: Path, sheet: str) -> list[dict[str, Any]]:
    book = load_workbook(workbook, read_only=True, data_only=True)
    worksheet = book[sheet]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    rows = [dict(zip(headers, values)) for values in iterator if any(value is not None for value in values)]
    book.close()
    return rows


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8") as handle:
        return [json.loads(line) for line in handle if line.strip()]


def display_change_type(raw: Any) -> str:
    value = clean(raw).lower()
    if value == "not_comparable":
        return "modified"
    if value not in DISPLAY_CHANGE_TYPES:
        raise ValueError(f"Unsupported change type: {value!r}")
    return value


def validation_pass(row: dict[str, Any]) -> bool:
    validation = parse_json(row.get("evidence_validation"), {})
    return bool(validation.get("required_evidence_valid"))


def parse_evidence(value: Any, *, verified: bool, method: str) -> list[dict[str, Any]]:
    payload = parse_json(value, [])
    if isinstance(payload, dict):
        payload = [payload]
    if not isinstance(payload, list):
        payload = [{"quote": clean(value), "page": None}]
    result: list[dict[str, Any]] = []
    seen: set[tuple[str, int | None]] = set()
    for entry in payload:
        if not isinstance(entry, dict):
            entry = {"quote": clean(entry), "page": None}
        quote = clean(entry.get("quote"))
        page = as_int(entry.get("page"), default=-1)
        page_value = None if page < 0 else page
        marker = (quote, page_value)
        if not quote or marker in seen:
            continue
        seen.add(marker)
        result.append(
            {
                "quote": quote,
                "page": page_value,
                "claimedPage": page_value,
                "verified": verified,
                "method": method,
            }
        )
    return result


def atomic_from_row(row: dict[str, Any], route: str, ordinal: int) -> dict[str, Any]:
    cross = route == "cross-period"
    raw_type = clean(row.get("change_type"))
    change_type = display_change_type(raw_type)
    verified = validation_pass(row) if cross else False
    method = "deepseek_postprocess_text_match" if verified else "deepseek_cleaned_export"
    score = as_int(row.get("score") if cross else row.get("score_numeric"))
    review_reason = clean(row.get("review_reason")) if cross else clean(row.get("manual_note"))
    return {
        "id": f"{clean(row.get('job_id'))}::{clean(row.get('clause_id')) or ordinal}",
        "clauseId": clean(row.get("clause_id")),
        "variableCode": clean(row.get("canonical_variable_code")),
        "variableFamily": clean(row.get("variable_family")),
        "variableName": clean(row.get("variable_name")),
        "subclauseLabel": clean(row.get("subclause_label")),
        "rawChangeType": raw_type,
        "changeType": change_type,
        "changeTypeLabel": CHANGE_TYPE_LABELS[change_type],
        "measurementType": clean(row.get("measurement_type")),
        "oldValue": clean(row.get("old_value")),
        "newValue": clean(row.get("new_value")),
        "numericDirection": clean(row.get("numeric_direction") if cross else row.get("final_numeric_direction")),
        "economicBurdenDirection": clean(row.get("economic_burden_direction") if cross else row.get("final_economic_burden_direction")),
        "controlShift": clean(row.get("control_shift") if cross else row.get("final_control_shift")),
        "performanceDirection": clean(row.get("performance_direction")),
        "actor": clean(row.get("actor")),
        "affectedParty": clean(row.get("affected_party")),
        "financingType": clean(row.get("financing_type")),
        "bindingStatus": clean(row.get("binding_status")),
        "oldEvidence": parse_evidence(row.get("old_evidence"), verified=verified, method=method),
        "newEvidence": parse_evidence(row.get("new_evidence"), verified=verified, method=method),
        "score": score,
        "reason": clean(row.get("score_reason")),
        "outcomeReady": as_bool(row.get("outcome_ready")) if cross else True,
        "needsReview": as_bool(row.get("review_required")) if cross else False,
        "reviewReason": review_reason,
        "inferenceLimit": clean(row.get("inference_limit")),
    }


def change_counts(atomics: Iterable[dict[str, Any]]) -> dict[str, int]:
    counts = Counter(atomic["changeType"] for atomic in atomics)
    return {key: int(counts.get(key, 0)) for key in DISPLAY_CHANGE_TYPES}


def page_range(evidence: Iterable[dict[str, Any]]) -> list[int | None]:
    pages = [entry["page"] for entry in evidence if entry.get("page") is not None]
    return [min(pages), max(pages)] if pages else [None, None]


def aggregate_evidence(atomics: Iterable[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[tuple[str, int | None]] = set()
    for atomic in atomics:
        for entry in atomic[key]:
            marker = (entry["quote"], entry["page"])
            if marker not in seen:
                seen.add(marker)
                result.append(entry)
    return result


def select_direction(atomics: list[dict[str, Any]]) -> str:
    for key in ("economicBurdenDirection", "controlShift", "numericDirection"):
        values = {clean(atomic.get(key)) for atomic in atomics} - IGNORED_DIRECTIONS
        if len(values) == 1:
            return next(iter(values))
        if len(values) > 1:
            return "mixed"
    values = {atomic["changeType"] for atomic in atomics}
    return next(iter(values)) if len(values) == 1 else "mixed"


def build_summary(atomics: list[dict[str, Any]]) -> str:
    ordered = sorted(atomics, key=lambda row: (-row["score"], row["variableName"]))
    snippets: list[str] = []
    for atomic in ordered:
        text = atomic["reason"] or atomic["variableName"]
        if text and text not in snippets:
            snippets.append(text)
        if len(snippets) == 2:
            break
    lead = f"{len(atomics)} atomic change{'s' if len(atomics) != 1 else ''}."
    return clean(f"{lead} {' '.join(snippets)}")


def pair_id(job_id: str, route: str) -> str:
    value = re.sub(rf"^fdd__{re.escape(route)}__", "", job_id)
    return re.sub(r"__item_\d+$", "", value)


def comparison_row(
    *,
    route: str,
    job_id: str,
    company: str,
    old_year: int,
    new_year: int,
    item: int,
    item_title: str,
    atomics: list[dict[str, Any]],
    summary: str,
    comparison_complete: bool,
    old_document: str = "",
    new_document: str = "",
    old_source: str = "Source not exported",
    new_source: str = "Source not exported",
    old_pages: list[int | None] | None = None,
    new_pages: list[int | None] | None = None,
    supplied_pair_id: str = "",
    stable: bool = False,
    needs_review: bool = False,
    review_reason: str = "",
    pipeline: str = "",
) -> dict[str, Any]:
    old_evidence = aggregate_evidence(atomics, "oldEvidence")
    new_evidence = aggregate_evidence(atomics, "newEvidence")
    score = max((atomic["score"] for atomic in atomics), default=0)
    contractual = any(
        atomic["score"] >= 4 and atomic["bindingStatus"] in CONTRACTUAL_BINDING
        for atomic in atomics
    )
    review_reasons = [review_reason] if review_reason else []
    for atomic in atomics:
        if atomic["reviewReason"] and atomic["reviewReason"] not in review_reasons:
            review_reasons.append(atomic["reviewReason"])
    return {
        "route": route,
        "id": job_id,
        "pairId": supplied_pair_id or pair_id(job_id, route),
        "company": company,
        "oldYear": old_year,
        "newYear": new_year,
        "yearGap": new_year - old_year,
        "oldSource": old_source,
        "newSource": new_source,
        "sameSource": old_source == new_source,
        "item": item,
        "itemTitle": item_title,
        "score": score,
        "substantive": any(atomic["score"] >= 3 for atomic in atomics),
        "contractual": contractual,
        "routine": False,
        "direction": select_direction(atomics) if atomics else "no_substantive_change",
        "summary": summary,
        "statedReason": "",
        "confidence": "outcome-ready" if any(atomic["outcomeReady"] for atomic in atomics) else "production candidate",
        "needsReview": needs_review or any(atomic["needsReview"] for atomic in atomics),
        "reviewReason": " ".join(review_reasons[:3]),
        "evidenceStatus": "deepseek_postprocess_validated" if route == "cross-period" else "deepseek_conservative_cleaned",
        "evidenceGatePass": not any(atomic["needsReview"] and "evidence" in atomic["reviewReason"].lower() for atomic in atomics),
        "oldDocument": old_document,
        "newDocument": new_document,
        "oldPages": old_pages or page_range(old_evidence),
        "newPages": new_pages or page_range(new_evidence),
        "oldEvidence": old_evidence,
        "newEvidence": new_evidence,
        "samplingWeight": None,
        "samplingStratum": "",
        "inSbaDirectory": False,
        "replacement": False,
        "replacementLevel": "",
        "financingGuardrailPass": item != 10 or all(atomic["financingType"] != "unsupported" for atomic in atomics),
        "financingWarnings": "",
        "scoringSource": "deepseek_api_cleaned" if route == "consecutive" else "deepseek_api_v4_5_production",
        "model": MODEL,
        "pipeline": pipeline,
        "comparisonComplete": comparison_complete,
        "stable": stable,
        "atomicChangeCount": len(atomics),
        "outcomeReadyAtomicChanges": sum(atomic["outcomeReady"] for atomic in atomics),
        "reviewRequiredAtomicChanges": sum(atomic["needsReview"] for atomic in atomics),
        "highImpactAtomicChanges": sum(atomic["score"] >= 4 for atomic in atomics),
        "changeTypeCounts": change_counts(atomics),
        "inferenceLimit": " ".join(dict.fromkeys(atomic["inferenceLimit"] for atomic in atomics if atomic["inferenceLimit"]))[:1800],
        "atomicChanges": atomics,
    }


def build_consecutive(workbook: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    atomic_source = [row for row in read_sheet_rows(workbook, "Clean Atomic") if clean(row.get("final_status")) == "INCLUDE_CONSERVATIVE"]
    item_source = read_sheet_rows(workbook, "Item Results")
    titles = {as_int(row["item"]): clean(row["item_title"]) for row in item_source}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in atomic_source:
        grouped[clean(row["job_id"])].append(row)

    rows: list[dict[str, Any]] = []
    for job_id, source_rows in sorted(grouped.items()):
        atomics = [atomic_from_row(row, "consecutive", index) for index, row in enumerate(source_rows)]
        first = source_rows[0]
        rows.append(
            comparison_row(
                route="consecutive",
                job_id=job_id,
                company=clean(first.get("company")),
                old_year=as_int(first.get("old_year")),
                new_year=as_int(first.get("new_year")),
                item=as_int(first.get("item")),
                item_title=titles[as_int(first.get("item"))],
                atomics=atomics,
                summary=build_summary(atomics),
                comparison_complete=True,
                pipeline="cleaned-v4.3",
            )
        )

    item_stats: list[dict[str, Any]] = []
    by_item = defaultdict(list)
    for row in rows:
        by_item[row["item"]].append(row)
    for source in item_source:
        item = as_int(source["item"])
        item_rows = by_item[item]
        atomics = [atomic for row in item_rows for atomic in row["atomicChanges"]]
        complete = as_int(source.get("complete_comparisons"))
        item_stats.append(
            {
                "item": item,
                "title": titles[item],
                "comparisons": complete,
                "incomplete": as_int(source.get("incomplete_comparisons")),
                "changeJobs": as_int(source.get("conservative_change_jobs")),
                "changeJobRate": round(float(source.get("conservative_change_rate") or 0) * 100, 1),
                "highImpactJobs": as_int(source.get("high_impact_jobs_score_4_5")),
                "highImpactRate": round(float(source.get("high_impact_rate") or 0) * 100, 1),
                "atomicChanges": len(atomics),
                "outcomeReadyAtomicChanges": len(atomics),
                "reviewRequiredAtomicChanges": 0,
                "changeTypeCounts": change_counts(atomics),
            }
        )

    all_atomics = [atomic for row in rows for atomic in row["atomicChanges"]]
    metadata = {
        "route": "consecutive",
        "design": "DeepSeek v4.3 conservative cleaned consecutive-year results",
        "model": MODEL,
        "preparedJobs": 4557,
        "exportedJobs": 4556,
        "completeComparisons": 4352,
        "incompleteComparisons": 204,
        "comparisons": len(rows),
        "includedChangeJobs": len(rows),
        "atomicChanges": len(all_atomics),
        "outcomeReadyAtomicChanges": len(all_atomics),
        "reviewRequiredAtomicChanges": 0,
        "highImpactChangeJobs": sum(row["score"] >= 4 for row in rows),
        "highImpactAtomicChanges": sum(atomic["score"] >= 4 for atomic in all_atomics),
        "uniquePairs": len({row["pairId"] for row in rows}),
        "items": list(range(1, 24)),
        "includesNoChange": False,
        "unit": "included_change_job_with_atomic_changes",
        "source": workbook.name,
        "changeTypeCounts": change_counts(all_atomics),
        "notComparableDisplayedAsModified": sum(atomic["rawChangeType"] == "not_comparable" for atomic in all_atomics),
    }
    return rows, item_stats, metadata


def build_cross_period(root: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    production = root / "production_cross_period_v4_2"
    input_jobs = read_jsonl(root / "production_cross_period_jobs_v4_2.jsonl")
    job_payloads = {entry["job_id"]: entry["payload"] for entry in input_jobs}
    summaries = read_csv_rows(production / "flat/fdd_item_summary.csv")
    atomic_source = read_csv_rows(production / "flat/fdd_atomic_changes.csv")
    result_records = read_jsonl(production / "results.jsonl")
    pipelines = {entry["job_id"]: clean(entry.get("pipeline_version")) for entry in result_records}
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for source in atomic_source:
        grouped[clean(source["job_id"])].append(source)

    rows: list[dict[str, Any]] = []
    for summary in summaries:
        job_id = clean(summary["job_id"])
        payload = job_payloads[job_id]
        old_doc = payload["old_document"]
        new_doc = payload["new_document"]
        atomics = [atomic_from_row(source, "cross-period", index) for index, source in enumerate(grouped.get(job_id, []))]
        old_extract = old_doc.get("extraction", {})
        new_extract = new_doc.get("extraction", {})
        rows.append(
            comparison_row(
                route="cross-period",
                job_id=job_id,
                company=clean(payload.get("company")),
                old_year=as_int(payload.get("old_year")),
                new_year=as_int(payload.get("new_year")),
                item=as_int(payload.get("item")),
                item_title=clean(payload.get("item_title")),
                atomics=atomics,
                summary=clean(summary.get("one_sentence_takeaway")) or build_summary(atomics),
                comparison_complete=as_bool(summary.get("comparison_complete")),
                old_document=clean(old_doc.get("filename")),
                new_document=clean(new_doc.get("filename")),
                old_source=clean(old_doc.get("source")),
                new_source=clean(new_doc.get("source")),
                old_pages=[as_int(old_extract.get("start_pdf_page"), -1), as_int(old_extract.get("end_pdf_page"), -1)],
                new_pages=[as_int(new_extract.get("start_pdf_page"), -1), as_int(new_extract.get("end_pdf_page"), -1)],
                supplied_pair_id=clean(payload.get("pair_id")),
                stable=as_bool(summary.get("stable")),
                needs_review=as_bool(summary.get("overall_review_required")),
                review_reason="",
                pipeline=pipelines.get(job_id, ""),
            )
        )

    for row in rows:
        row["oldPages"] = [None if value < 0 else value for value in row["oldPages"]]
        row["newPages"] = [None if value < 0 else value for value in row["newPages"]]

    by_item = defaultdict(list)
    for row in rows:
        by_item[row["item"]].append(row)
    item_stats: list[dict[str, Any]] = []
    for item in sorted(by_item):
        item_rows = by_item[item]
        atomics = [atomic for row in item_rows for atomic in row["atomicChanges"]]
        ready_jobs = sum(any(atomic["outcomeReady"] for atomic in row["atomicChanges"]) for row in item_rows)
        high_ready_jobs = sum(any(atomic["outcomeReady"] and atomic["score"] >= 4 for atomic in row["atomicChanges"]) for row in item_rows)
        detected_jobs = sum(bool(row["atomicChanges"]) for row in item_rows)
        ready_atomics = [atomic for atomic in atomics if atomic["outcomeReady"]]
        item_stats.append(
            {
                "item": item,
                "title": item_rows[0]["itemTitle"],
                "comparisons": len(item_rows),
                "incomplete": sum(not row["comparisonComplete"] for row in item_rows),
                "detectedChangeJobs": detected_jobs,
                "detectedChangeJobRate": round(detected_jobs / len(item_rows) * 100, 1),
                "changeJobs": ready_jobs,
                "changeJobRate": round(ready_jobs / len(item_rows) * 100, 1),
                "highImpactJobs": high_ready_jobs,
                "highImpactRate": round(high_ready_jobs / len(item_rows) * 100, 1),
                "atomicChanges": len(atomics),
                "outcomeReadyAtomicChanges": len(ready_atomics),
                "reviewRequiredAtomicChanges": sum(atomic["needsReview"] for atomic in atomics),
                "changeTypeCounts": change_counts(atomics),
                "outcomeReadyChangeTypeCounts": change_counts(ready_atomics),
            }
        )

    audit = json.loads((production / "audit/audit_report.json").read_text(encoding="utf-8"))
    all_atomics = [atomic for row in rows for atomic in row["atomicChanges"]]
    metadata = {
        "route": "cross-period",
        "design": "DeepSeek v4.5 final cross-period production package",
        "model": MODEL,
        "targetComparisons": 1350,
        "preparedJobs": len(input_jobs),
        "unresolvedInputs": 16,
        "successfulComparisons": len(rows),
        "completeComparisons": sum(row["comparisonComplete"] for row in rows),
        "comparisons": len(rows),
        "atomicChanges": len(all_atomics),
        "detectedChangeJobs": sum(bool(row["atomicChanges"]) for row in rows),
        "outcomeReadyChangeJobs": sum(any(atomic["outcomeReady"] for atomic in row["atomicChanges"]) for row in rows),
        "highImpactReadyChangeJobs": sum(any(atomic["outcomeReady"] and atomic["score"] >= 4 for atomic in row["atomicChanges"]) for row in rows),
        "outcomeReadyAtomicChanges": sum(atomic["outcomeReady"] for atomic in all_atomics),
        "reviewRequiredAtomicChanges": sum(atomic["needsReview"] for atomic in all_atomics),
        "uniquePairs": len({row["pairId"] for row in rows}),
        "items": sorted(by_item),
        "includesNoChange": True,
        "productionApproved": bool(audit.get("production_approved")),
        "analysisReady": bool(audit.get("analysis_ready")),
        "auditDecision": clean(audit.get("decision")),
        "auditNote": clean(audit.get("production_approval_note")),
        "source": "FDD_cross_period_v4_5_upload.zip",
        "changeTypeCounts": change_counts(all_atomics),
        "notComparableDisplayedAsModified": sum(atomic["rawChangeType"] == "not_comparable" for atomic in all_atomics),
    }
    return rows, item_stats, metadata


def compact_case(row: dict[str, Any]) -> dict[str, Any]:
    excluded = {"atomicChanges", "oldEvidence", "newEvidence", "inferenceLimit"}
    return {key: value for key, value in row.items() if key not in excluded}


def write_chunks(directory: Path, prefix: str, rows: list[dict[str, Any]]) -> list[str]:
    chunks: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    current_bytes = 24
    for row in rows:
        row_bytes = len(json.dumps(row, ensure_ascii=False, separators=(",", ":")).encode("utf-8")) + 1
        if current and current_bytes + row_bytes > MAX_SHARD_BYTES:
            chunks.append(current)
            current = []
            current_bytes = 24
        current.append(row)
        current_bytes += row_bytes
    if current or not chunks:
        chunks.append(current)
    filenames: list[str] = []
    for index, chunk in enumerate(chunks, start=1):
        filename = f"{prefix}.json" if index == 1 else f"{prefix}-{index:02d}.json"
        (directory / filename).write_text(
            json.dumps({"rows": chunk}, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        filenames.append(filename)
    return filenames


def write_route(site_root: Path, route: str, rows: list[dict[str, Any]], item_stats: list[dict[str, Any]], metadata: dict[str, Any]) -> None:
    directory = site_root / "public/data" / route
    directory.mkdir(parents=True, exist_ok=True)
    for old_file in directory.glob("*.json"):
        old_file.unlink()
    files: list[str] = []
    for item in metadata["items"]:
        item_rows = [row for row in rows if row["item"] == item]
        files.extend(write_chunks(directory, f"item-{item:02d}", item_rows))
    case_files = write_chunks(directory, "cases", [compact_case(row) for row in rows])
    metadata["rowsPerItem"] = {str(item): sum(row["item"] == item for row in rows) for item in metadata["items"]}
    index = {
        "metadata": metadata,
        "changeTypeLabels": CHANGE_TYPE_LABELS,
        "itemStats": item_stats,
        "files": files,
        "caseFiles": case_files,
    }
    (directory / "index.json").write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")


def generated_item(row: dict[str, Any], rank: int) -> dict[str, Any]:
    return {
        "rank": rank,
        "item": row["item"],
        "title": row["title"],
        "share": row["changeJobRate"],
        "major": row["highImpactRate"],
        "n": row["comparisons"],
        "incomplete": row["incomplete"],
        "atomicChanges": row["atomicChanges"],
        "outcomeReadyAtomicChanges": row["outcomeReadyAtomicChanges"],
        "reviewRequiredAtomicChanges": row["reviewRequiredAtomicChanges"],
        "detectedShare": row.get("detectedChangeJobRate"),
        "changeTypes": row["changeTypeCounts"],
    }


def write_generated_results(site_root: Path, consecutive_stats: list[dict[str, Any]], cross_stats: list[dict[str, Any]], consecutive_meta: dict[str, Any], cross_meta: dict[str, Any]) -> None:
    consecutive_ordered = sorted(consecutive_stats, key=lambda row: (-row["changeJobRate"], row["item"]))
    cross_ordered = sorted(cross_stats, key=lambda row: (-row["changeJobRate"], row["item"]))
    payload = {
        "consecutiveItemsGenerated": [generated_item(row, index) for index, row in enumerate(consecutive_ordered, 1)],
        "crossPeriodItemsGenerated": [generated_item(row, index) for index, row in enumerate(cross_ordered, 1)],
        "routeSummariesGenerated": {
            "consecutive": {"metadata": consecutive_meta, "items": consecutive_stats},
            "cross-period": {"metadata": cross_meta, "items": cross_stats},
        },
    }
    lines = ["// Generated by scripts/generate_deepseek_site_data.py. Do not hand-edit."]
    for key, value in payload.items():
        lines.append(f"export const {key} = {json.dumps(value, ensure_ascii=False, indent=2)};")
    (site_root / "app/generated-results.ts").write_text("\n\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    args = parse_args()
    consecutive_rows, consecutive_stats, consecutive_meta = build_consecutive(args.consecutive_workbook)
    cross_rows, cross_stats, cross_meta = build_cross_period(args.cross_package_dir)

    assert len(consecutive_rows) == 1661
    assert consecutive_meta["atomicChanges"] == 7705
    assert sum(consecutive_meta["changeTypeCounts"].values()) == 7705
    assert len(cross_rows) == 1334
    assert cross_meta["atomicChanges"] == 14505
    assert cross_meta["outcomeReadyAtomicChanges"] == 6211
    assert cross_meta["reviewRequiredAtomicChanges"] == 6934
    assert sum(cross_meta["changeTypeCounts"].values()) == 14505

    write_route(args.site_root, "consecutive", consecutive_rows, consecutive_stats, consecutive_meta)
    write_route(args.site_root, "cross-period", cross_rows, cross_stats, cross_meta)
    write_generated_results(args.site_root, consecutive_stats, cross_stats, consecutive_meta, cross_meta)

    print(json.dumps({"consecutive": consecutive_meta, "cross-period": cross_meta}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
